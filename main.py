import os
import openpyxl
from openpyxl.workbook import Workbook

from keshiSheet import keshisheet
from keshi import keshi
from wanzixi21 import wanzixi21
from wanzixi2223 import wanzixi2223
from wanzixiSheet import wanzixisheet
from zaodu import zaodu
from zaoduSheet import zaodusheet


def unmerge_and_fill_cells():
    workbook = openpyxl.load_workbook(filePath)  # 返回一个workbook数据类型的值
    sheet = workbook['课表']

    all_merged_cell_ranges = list(
        sheet.merged_cells.ranges
    )
    # print(all_merged_cell_ranges)
    for merged_cell_range in all_merged_cell_ranges:
        merged_cell = merged_cell_range.start_cell
        sheet.unmerge_cells(range_string=merged_cell_range.coord)

        for row_index, col_index in merged_cell_range.cells:
            cell = sheet.cell(row=row_index, column=col_index)
            cell.value = merged_cell.value
    workbook.save("temp.xlsx")
    workbook.close()


class ClassTime:
    def __init__(self, name, classname, time, classtype, ismore):
        self.name = name
        self.time = time
        self.classname = classname
        self.classtype = classtype
        self.ismore = ismore
    name = None
    classname = None  # 班级
    time = None  # 日期
    classtype = None  # 1代表早读  2代表常规课程  3代表晚自习
    ismore = False  # True代表该课时是1.5倍课时


def getClassTime():
    wb = openpyxl.load_workbook(processPath)
    ws = wb['课表']
    # print(ws["BO30"].value.type())

    # 设置起始行列索引
    start_row = 4
    start_col = 2

    # 获取结束行列索引
    end_row = ws.max_row
    end_col = ws.max_column

    print(end_row)
    print(end_col)
    # 遍历这个区间内的所有格子
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        # 迭代每行中的单元格
        for cell in row:
            # 将单元格存入classTimeObject中
            # print(cell.coordinate, cell.value)
            row = cell.row
            col = cell.column
            name = None
            classname = None
            ismore = False
            classtype = None
            if cell.value == None:
                continue

            classname = ws.cell(row=row, column=1).value
            # 判断是否结尾有（1.5）课时，并完成isMore和name的定义
            strnow = cell.value
            if "1.5" in strnow:
                ismore = True
                name = strnow[:-5]
            else:
                ismore = False
                name = strnow

            # 判断是否为早自习，晚自习
            strnow = ws.cell(row=3, column=col).value
            # if strnow == None:
            #     print(row)
            #     print(col)
            if "早自习" in strnow or "早读" in strnow:
                classtype = 1
            elif "晚自习" in strnow:
                classtype = 3
            else:
                classtype = 2

            time = ws.cell(row=1, column=col).value

            # 将对象加入到列表中
            classTimeObject.append(ClassTime(name, classname, time, classtype, ismore))
    wb.close()

def getPersonSheet():
    wb = openpyxl.load_workbook(processPath)
    ws = wb['名单表']
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            personSheet.append(cell.value)
    wb.close

if __name__ == "__main__":
    path = input("请输入文件名:")
    # os.chdir(path)  # 修改工作路径
    filePath = path
    processPath = r"temp.xlsx"
    unmerge_and_fill_cells() #转化课表的多行数据填充每一行
    classTimeObject = []
    personSheet = []
    getClassTime()
    getPersonSheet()
    cnt = 0

    wb = Workbook()

    # 生成课时表
    keshisheet(classTimeObject, wb)
    keshi(classTimeObject, personSheet, wb)

    zaodusheet(classTimeObject, wb)
    zaodu(classTimeObject, personSheet, wb)

    wanzixisheet(classTimeObject, wb)
    wanzixi21(classTimeObject, personSheet, wb)
    wanzixi2223(classTimeObject, personSheet, wb)

    wb.save("result.xlsx")
    wb.close()

    os.remove("temp.xlsx")



